from openpyxl import load_workbook

wb1 = load_workbook('wazirx.xlsx')
sheet1 = wb1['Sheet1']
total_wazir_row = sheet1.max_row

wb2 = load_workbook('crypto_report.xlsx')
sheet2 = wb2['Sheet1']
total_binance_row = sheet2.max_row

wb3 = load_workbook('cryptosampl.xlsx')
sheet3 = wb3['Sheet1']

track_row = 2
for row in range(2, total_wazir_row+1):
    coin = sheet1.cell(row= row, column=1).value.lower().strip()
    for row1 in range(2, total_binance_row+1):
        if coin == sheet2.cell(row= row1, column=1).value.lower().strip():
            sheet3.cell(row = track_row, column=1).value = coin
            sheet3.cell(row = track_row, column=2).value = sheet2.cell(row=row1, column=2).value
            sheet3.cell(row = track_row, column=3).value = sheet1.cell(row=row, column=2).value
            sheet3.cell(row = track_row, column=4).value = sheet1.cell(row=row, column=3).value
            track_row += 1
    wb3.save('final_crypto_report.xlsx')




