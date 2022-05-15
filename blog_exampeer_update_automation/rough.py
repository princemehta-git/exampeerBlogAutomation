# from selenium import webdriver
# from webdriver_manager.chrome import ChromeDriverManager
# from selenium.webdriver.support.select import Select
# import docx
# import glob
# import sys
# import os
# from time import sleep
# import sys
# import pyperclip
# from selenium.webdriver.common.action_chains import ActionChains
# from selenium.webdriver.common.keys import Keys
# from googlesearch import search
# from PIL import Image
# import time
#
# options = webdriver.ChromeOptions()
# # options.headless = True
# user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.83 Safari/537.36"
# options.add_argument(f'user-agent={user_agent}')
# options.add_argument("--window-size=1920,1080")
# options.add_argument('--ignore-certificate-errors')
# options.add_argument('--allow-running-insecure-content')
# options.add_argument("--disable-extensions")
# options.add_argument("--proxy-server='direct://'")
# options.add_argument("--proxy-bypass-list=*")
# options.add_argument("--start-maximized")
# options.add_argument('--disable-gpu')
# options.add_argument('--disable-dev-shm-usage')
# options.add_argument('--no-sandbox')
# options.add_argument("--app=https://www.google.com")
# options.add_experimental_option("useAutomationExtension", False)
# options.add_experimental_option("excludeSwitches", ["enable-automation"])
#
#
# driver = webdriver.Chrome(ChromeDriverManager().install(), options = options)
# driver.implicitly_wait(5)
# login = False
# while not login:
#     driver.get('https://currentaffairsnews.in/admin-login')
#     driver.find_element_by_id('input-email').send_keys('admin@gmail.com')
#     driver.find_element_by_id('input-password').send_keys('admin')
#     driver.find_element_by_id('btn-login').click()
#     try:
#         driver.find_elements_by_xpath('//a[@class="breadcrumb--active"]')
#         login = True
#     except:
#         login = False
# # sleep(3)
#
#
# path = r"C:\Users\PRINCE MEHTA\Dropbox\CURRENT AFFAIRS\March 2022\Daily Updates\14 Mar"
# docx_files = glob.glob(path + "\\*.docx")
# # image files must be all in jpg
# img_files = glob.glob(path + "\\*.jpg")
# if len(img_files) != 10:
#     rest_img_files = glob.glob(path + '\\*.png') + glob.glob(path + '\\*.svg') + glob.glob(path + '\\*.webp') + glob.glob(path + '\\*.jpeg') + glob.glob(path + '\\*.jfif') + glob.glob(path + '\\*.pjpeg') + glob.glob(path + '\\*.pjp')
#     for img in rest_img_files:
#         Image.open(img).convert("RGB").save(img.replace(img.split('.')[-1], 'jpg'))
#
# img_files = glob.glob(path + "\\*.jpg")
#
#
# categories_eng = {}
#
# categories_hin = {}
#
#
# # category_values = {'Agreements':"53",
# #                'Appointments':"54",
# #                'Awards':"55",
# #                'Banking':"56",
# #                'Books and Authors':"57",
# #                'Defence':"58",
# #                'Economy':"59",
# #                'Important Days':"60",
# #                'International':"61",
# #                'Miscellaneous':"62",
# #                'National':"63",
# #                'Obituaries':"64",
# #                'Ranks & Reports':"65",
# #                'Schemes and Committees':"66",
# #                'Science & Technology':"67",
# #                'Sports':"68",
# #                'States':"69",
# #                'Summits and Conference':"70"}
#
# topics = ['Agreements', 'Appointments', 'Awards', 'Banking', 'Books and Authors', 'Defence', 'Economy', 'Important Days', 'International', 'Miscellaneous', 'National', 'Obituaries', 'Ranks & Reports', 'Schemes and Committees', 'Science & Technology', 'Sports', 'States', 'Summits and Conference']
#
# if len(docx_files) == 2:
#     for file in docx_files[:-1]:
#         if 'eng' in file.lower():
#             paras = docx.Document(file).paragraphs
#             # updating topics/categories
#             for para_index, para in enumerate(paras):
#                 if 'daily update' in para.text.lower():
#                     categories_eng[para_index]=para.text.split(':')[-1].strip()
#             categories_enlis = list(categories_eng.keys())
#             # print('categories_eng:::', categories_eng)
#             # print('categories_enlist::::::', categories_enlis)
#             if len(categories_eng) == 10:
#                 for count, category_index in enumerate(categories_eng):
#                     driver.get('https://currentaffairsnews.in/add-blog/side-menu/light')
#
#                     # choosing language
#                     driver.find_elements_by_xpath('//div[@class="select-label tail-select-container"]')[0].click()
#                     # sleep(3)
#                     lan_opts = driver.find_elements_by_xpath('//ul[@class="dropdown-optgroup"]')[2].find_elements_by_tag_name('li')
#                     for lan_opt in lan_opts:
#                         if 'english' in lan_opt.text.lower():
#                             lan_opt.click()
#
#                     # for choosing topic
#                     # topic = False
#                     for topic in topics:
#                         # # print(categories_eng[category_index][0:4], topic)
#                         if categories_eng[category_index][:4].lower() in topic.lower():
#                             break
#                         else:
#                             topic = 'Miscellaneous'
#
#                     driver.find_elements_by_xpath('//div[@class="select-label"]')[1].click()
#
#                     cat_opts = driver.find_elements_by_xpath('//ul[@class="dropdown-optgroup"]')[1].find_elements_by_tag_name('li')
#                     for cat_opt in cat_opts:
#                         # # print(cat_opt.text, topic)
#                         if topic.lower() in cat_opt.text.lower():
#                             cat_opt.click()
#                             # cat_opt.click()
#                             # print('matched and clicked')
#
#                     # adding title:
#                     tit = None
#                     if ((paras[categories_enlis[count]+1].text.strip() != '' ) and (paras[categories_enlis[count]+1].text.strip() != None)):
#                         title = paras[categories_enlis[count]+1].text.strip(':-')
#                         driver.find_element_by_xpath('//input[@name="title"]').click()
#
#                         past = pyperclip.paste()
#                         pyperclip.copy(title)
#                         ActionChains(driver).send_keys(Keys.CONTROL, 'v')
#                         pyperclip.copy(past)
#                         tit = 1
#                         # print('choosing tit:::', title, 'of para index::',categories_enlis[count]+1 )
#                     elif (paras[categories_enlis[count]+2].text.strip().endswith(':-')) or ((paras[categories_enlis[count]+1].text.strip() != '' ) and (paras[categories_enlis[count]+1].text.strip() != None)):
#                         title = paras[categories_enlis[count] + 2].text.strip(':-')
#                         driver.find_element_by_xpath('//input[@name="title"]').click()
#
#                         past = pyperclip.paste()
#                         pyperclip.copy(title)
#                         ActionChains(driver).key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
#                         pyperclip.copy(past)
#                         tit = 2
#
#                     # print('title :::::', title, tit)
#
#
#                     # adding description
#                     if paras[categories_enlis[count]+2].text.strip() != '' and paras[categories_enlis[count]+2].text.strip() != None and tit != 2:
#                         discription = paras[categories_enlis[count] + 2].text
#                         driver.find_element_by_xpath('//iframe[@class="cke_wysiwyg_frame cke_reset"]').click()
#
#                         past = pyperclip.paste()
#                         pyperclip.copy(discription)
#                         ActionChains(driver).key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
#                         pyperclip.copy(past)
#                     elif paras[categories_enlis[count]+3].text.strip() != '' and paras[categories_enlis[count]+3].text.strip() != None:
#                         discription = paras[categories_enlis[count] + 3].text
#                         driver.find_element_by_xpath('//iframe[@class="cke_wysiwyg_frame cke_reset"]').click()
#
#                         past = pyperclip.paste()
#                         pyperclip.copy(discription)
#                         ActionChains(driver).key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
#                         pyperclip.copy(past)
#                     elif paras[categories_enlis[count]+4].text.strip() != '' and paras[categories_enlis[count]+4].text.strip() != None:
#                         discription = paras[categories_enlis[count] + 4].text
#                         driver.find_element_by_xpath('//iframe[@class="cke_wysiwyg_frame cke_reset"]').click()
#
#                         past = pyperclip.paste()
#                         pyperclip.copy(discription)
#                         ActionChains(driver).key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
#                         pyperclip.copy(past)
#
#                     # print('discription :::::', discription)
#
#                     # adding url
#                     link = search(title, tld="co.in", num=1, start=1, stop=1, pause=1.5, lang='english').__next__()
#                     driver.find_elements_by_xpath('//input[@name="url"]')[0].click()
#
#                     past = pyperclip.paste()
#                     pyperclip.copy(link)
#                     ActionChains(driver).key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
#                     pyperclip.copy(past)
#
#                     # print('link :::::', link)
#
#                     # scheduling
#                     currentTime = time.strftime("%H:%M")
#                     js = 'document.getElementsByName("schedule_time")[0].value = "%s";'%str(currentTime)
#                     driver.execute_script(js)
#
#                     # print('time_scheduler :::::', currentTime)
#
#                     # uploading image
#                     for img in img_files:
#                         if str(count+1).lower() in img.split('\\')[-1].lower():
#                             driver.find_elements_by_xpath('//input[@id="image"]')[0].send_keys(img)
#
#                             # print('img :::::', img.split('\\')[-1])
#                             break
#
#
#
#                     # clicking visibility options
#                     for opt in driver.find_elements_by_xpath('//input[@class="input border mr-2"]'):
#                         opt.click()
#
#
#
#                     sleep(20)
#
#
#
#
#         # elif 'hin' in file.lower():
#         #     paras = docx.Document(file).paragraphs
#         #     for para_index, para in enumerate(paras):
#         #         if 'डेली अपडेट' in para.text.lower():
#         #             categories_hin[para_index]=para.text.split(':')[-1].strip()
#
#
#
#




from time import sleep
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
# from bs4 import BeautifulSoup
from openpyxl import load_workbook


options = webdriver.ChromeOptions()



user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.83 Safari/537.36"
options.add_argument(f'user-agent={user_agent}')
options.add_argument("--window-size=1920,1080")
options.add_argument('--ignore-certificate-errors')
options.add_argument('--allow-running-insecure-content')
options.add_argument("--disable-extensions")
options.add_argument("--proxy-server='direct://'")
options.add_argument("--proxy-bypass-list=*")
options.add_argument("--start-maximized")
options.add_argument('--disable-gpu')
options.add_argument('--disable-dev-shm-usage')
options.add_argument('--no-sandbox')
options.add_argument("--app=https://www.google.com")
options.add_experimental_option("useAutomationExtension", False)
options.add_experimental_option("excludeSwitches", ["enable-automation"])
options.headless = True
#
#
# driver_wazir = webdriver.Chrome(ChromeDriverManager().install(), options=options)
# driver_wazir.implicitly_wait(3)
# driver_wazir.maximize_window()
# driver_wazir.get('https://wazirx.com/exchange/BTC-INR')
# driver_wazir.execute_script("document.querySelector('#root > div > div.sc-fvhFOF.eZrolp > div > div.sc-bdfBQB.sc-ksXiFW.hsXoKH.gRMbfd > div > div.sc-bdfBQB.sc-ikPAEB.eOlgOh.kgJOyC > div.sc-bdfBQB.sc-biBsmb.hsXoKH.gHzZfM > div > label:nth-child(2)').click();")


driver_binance = webdriver.Chrome(ChromeDriverManager().install(), options=options)
driver_binance.implicitly_wait(3)
driver_binance.maximize_window()
driver_binance.get('https://www.binance.com/en/markets/coinInfo')


binance_dict = {}
# wazir_dict_usdt = {}
# wazir_dict_inr = {}

 # binance
for main_idex in range(8):
    page = driver_binance.find_elements_by_class_name('css-hlqxzb')
    rows = driver_binance.find_elements_by_class_name('css-leyy1t')
    for index,row in enumerate(rows):
        try:
            key = rows[index].find_elements_by_class_name('css-1wp9rgv')[0].text
            # print(key,end=' ')
            value = rows[index].find_elements_by_class_name('css-ydcgk2')[0].find_element_by_css_selector('div').text
            # print(value)
            binance_dict[key] = value.strip('$')
        except:
            pass

        # if main_idex <7:
        #     driver_binance.execute_script("document.getElementsByClassName('css-hlqxzb')[%s].click();"%main_idex)
    if main_idex < 7:
        page[main_idex].click()
    # print('page:: ', len(binance_dict))

# wazir
# wazir_keys = ['zil', 'btc', 'eth', 'ftm', 'wrx', 'gto', 'shib', 'sol', 'iost', 'celr', 'trx', 'iotx', 'xrp', 'win', 'xvg', 'jasmy', 'jst', 'matic', 'rune', 'sc', 'coti', 'ooki', 'sand', 'ada', 'ctsi', 'front', 'tfuel', 'dock', 'one', 'gala', 'bnb', 'dent', 'vet', 'bake', 'bnt', 'dot', 'slp', 'nas', 'hot', 'doge', 'waves', 'bttc', 'stpt', 'stmx', 'cream', 'celo', 'luna', 'sun', 'alpaca', 'xlm', 'icp', 'enj', 'chr', 'ark', 'link', 'aave', 'gxs', 'mdx', 'rsr', 'mana', 'theta', 'dgb', 'sushi', 'omg', 'avax', 'busd', 'usdc', 'kmd', 'lrc', 'skl', 'dusk', '1inch', 'ape', 'reef', 'icx', 'ar', 'stx', 'mbox', 'bico', 'sxp', 'tko', 'hnt', 'xec', 'ftt', 'tlm', 'hbar', 'chz', 'ltc', 'yfii', 'axs', 'ogn', 'qnt', 'cos', 'dydx', 'near', 'cake', 'wtc', 'voxel', 'dodo', 'glmr', 'snx', 'dexe', 'dar', 'flux', 'ocean', 'rose', 'xtz', 'santos', 'push', 'atom', 't', 'ksm', 'grs', 'knc', 'ens', 'audio', 'ach', 'ren', 'grt', 'fun', 'etc', 'blz', 'qkc', 'aion', 'alice', 'band', 'ankr', 'eos', 'bat', 'comp', 'beta', 'req', 'mina', 'algo', 'kava', 'fil', 'qi', 'bch', 'brd', 'nbs', 'zrx', 'pyr', 'ctxc', 'bal', 'alpha', 'srm', 'astr', 'cocos', 'ckb', 'uni', 'xvs', 'uft', 'fet', 'key', 'waxp', 'egld', 'cfx', 'spell', 'ava', 'sys', 'rare', 'tomo', 'uma', 'super', 'adx', 'rvn', 'amp', 'iota', 'strax', 'bts', 'ez', 'dash', 'storj', 'alpine', 'data', 'mtl', 'pundix', 'nkn', 'crv', 'dego', 'trb', 'ata', 'burger', 'oxt', 'quick', 'flow', 'mir', 'ray', 'zec', 'pha', 'inj', 'mask', 'tusd', 'vgx', 'steem', 'neo', 'agld', 'mft', 'cvc', 'yfi', 'rad', 'poly', 'vite', 'fida', 'anc', 'rlc', 'lsk', 'firo', 'xno', 'btg', 'xem', 'snt', 'qtum', 'scrt', 'loom', 'bchsv', 'clv', 'hive', 'ant', 'imx', 'ardr', 'vib', 'lazio', 'dnt', 'api3', 'beam', 'ont', 'pnt', 'xmr', 'auction', 'klay', 'glm', 'mkr', 'porto', 'gno', 'chess', 'nuls', 'dcr', 'powr', 'rep', 'perp', 'ilv', 'nmr', 'paxg', 'idex', 'usdp', 'mln']
# sleep(3)
# # print(driver_wazir.execute_script("document.getElementsByClassName('ticker-item').length;"))
# # length = int(driver_wazir.execute_script("document.getElementsByClassName('ticker-item').length;"))
# elements = driver_wazir.find_elements_by_class_name('ticker-item')
# # print('elemnts: ', elements)
# for element in elements:
#
#     # wazir_dict_usdt[key]= driver_wazir.find_element_by_id('ticker-%s'%key).find_element_by_class_name('price-text ticker-price').text
#     # wazir_dict_inr[key]= driver_wazir.find_element_by_id('ticker-%s'%key).find_element_by_class_name('price-subtext').text
#
#     # scrape = BeautifulSoup(driver_wazir.find_element_by_id('ticker-%s'%key), 'html.parser').get_text()
#     # print(scrape)
#     # key = str(driver_wazir.execute_script("document.getElementsByClassName('ticker-item tablescraper-selected-row')[%s].getElementsByClassName('market-name-text')[0].textContent;"%index)).strip('/usdt').strip()
#
#     key = str(element.find_element_by_class_name('market-name-text').text).strip().strip('/usdt').strip('/USDT')
#
#     # print('key: ', key, end=' ')
#     # wazir_dict_usdt[key] = str(driver_wazir.execute_script("document.getElementsByClassName('ticker-item tablescraper-selected-row')[%s].getElementsByClassName('price-text ticker-price')[0].textContent;"%index)).strip(' USDT').strip()
#
#     wazir_dict_usdt[key] = str(element.find_element_by_css_selector('div:nth-child(3) > div:nth-child(1) > span').text).strip(' USDT').strip()
#
#     # print(wazir_dict_usdt[key])
#
#     # print('usdt: ', element.find_element_by_class_name('price-text ticker-price').text.strip(' USDT'), end=' ')
#     # wazir_dict_inr[key] = str(driver_wazir.execute_script("document.getElementsByClassName('ticker-item tablescraper-selected-row')[%s].getElementsByClassName('price-subtext')[0].textContent;"%index)).strip('₹').strip()
#     #
#     wazir_dict_inr[key] = str(element.find_element_by_css_selector('div:nth-child(3) > div:nth-child(2) > span').text).strip('₹').strip()
#     sleep(1)
#     # print(wazir_dict_inr[key])





print('binance::%s   '%len(binance_dict), binance_dict)
# print('------------------------------------------------------')
# print('wazir::   ', len(wazir_dict_usdt), wazir_dict_usdt)
# print('wazir::   ', len(wazir_dict_inr), wazir_dict_inr)

wb = load_workbook('cryptosampl.xlsx')

sheet = wb['Sheet1']
track_row = 2

# for key in wazir_dict_usdt:
#     if key in binance_dict.keys():
for key in binance_dict:
    sheet.cell(row=track_row, column=1).value = key
    sheet.cell(row=track_row, column=2).value = binance_dict[key]
    # sheet.cell(row=track_row, column=3).value = wazir_dict_usdt[key]
    # sheet.cell(row=track_row, column=4).value = wazir_dict_inr[key]
    track_row += 1

wb.save('crypto_report.xlsx')